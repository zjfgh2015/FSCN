import os
import random

import openpyxl
import pyexcel

def grab_all_data(cfg):
    dirs = os.listdir("./companies/")
    label_workload = openpyxl.load_workbook("labels.xlsx")["Sheet1"]
    company_name = label_workload["B2:B840"]
    company_label = label_workload["C2:C840"]
    all_company = []
    for company_path in dirs:
        if company_path[-1:] == "s":
            dest_file_name = "./companies/" + company_path[:-3] + "xlsx"
            pyexcel.save_book_as(file_name="./companies/" + company_path,
                                 dest_file_name=dest_file_name)
        else:
            workbook = openpyxl.load_workbook("./companies/" + company_path)
            try:
                sheet_new = workbook['new']
                sheet_hot = workbook['hot']
                row_new = sheet_new.max_row
                row_hot = sheet_hot.max_row
                content_new = sheet_new["B2:B" + str(row_new)]
                content_hot = sheet_hot["B2:B" + str(row_hot)]
            except Exception as e:
                print(e)
        for i, name in enumerate(company_name):
            if name[0].value in company_path or company_path[:-6] in name[0].value:
                label = company_label[i][0].value
                name_target = name[0].value
        content_new = [content[0].value for content in content_new]
        content_hot = [content[0].value for content in content_hot]

        for i in range(cfg.aug_num):
            new_remove = random.sample(range(0, len(content_new)), int(len(content_new) / 3))
            content_new = [content_new[i] for i in range(len(content_new)) if i not in new_remove]

            hot_remove = random.sample(range(0, len(content_hot)), int(len(content_hot) / 3))
            content_new = [content_hot[i] for i in range(len(content_hot)) if i not in hot_remove]
            object = {'company_name': name_target, "label": label, "content_new": content_new,
                      "content_hot": content_hot}
            all_company.append(object)

    return all_company


if __name__ == '__main__':
    al_data = grab_all_data()
    print()
